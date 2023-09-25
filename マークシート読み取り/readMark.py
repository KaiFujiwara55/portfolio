import openpyxl
import glob
import math
from matplotlib import pyplot as plt
import cv2
import numpy as np

# imgとmarkerの一致度の高い箇所を抽出, 返り値は[y, x]の座標
def match(img, marker):
    res = cv2.matchTemplate(img, marker, cv2.TM_CCOEFF_NORMED)
    loc = np.where(res > 0.7)

    # マーカーをwidthを基に分類
    loc_xy = []
    for x, y in zip(loc[1], loc[0]):
        loc_xy.append([x, y])
    loc_xy = sorted(loc_xy)
    markers = [[loc_xy[0]]]
    for data_x, data_y in loc_xy:
        data_x_b = markers[-1][-1][0]
        if (data_x - data_x_b) == 0:
            pass
        elif (data_x - data_x_b) > 10:
            markers.append([[data_x, data_y]])
        else:
            markers[-1].append([data_x, data_y])

    # マーカーを一点抽出
    res_max = 0
    res_maxs = []

    for datas in markers:
        for data in datas:
            res_max = res[data[1]][data[0]]
            if res[data[1]][data[0]] > res_max:
                res_max = loc[data[1]][data[0]]
        res_maxs.append([data[1], data[0]])

    return (res_maxs)

# 回転処理
def rotate(mks, marker, res_maxs):
    back = True
    if not (600 < res_maxs[1][0]) & (res_maxs[1][0] < 1200):
        mks = cv2.rotate(mks, cv2.ROTATE_180)
        res_maxs = match(mks, marker)
        back = False
    return (mks, res_maxs, back)

# 傾き補正
def tilt(mks, res_maxs, back=0):
    if back:
        x = res_maxs[8][1] - res_maxs[0][1]
        y = res_maxs[8][0] - res_maxs[0][0]
        degree = math.degrees(math.atan2(y, x))
        mat = cv2.getRotationMatrix2D(
            (res_maxs[1][1], res_maxs[1][0]), degree, 1.0)
        mks = cv2.warpAffine(mks, mat, (mks.shape[1], mks.shape[0]))
    else:
        # 3番目と11番目のマーカーの角度から傾きを補正
        x = res_maxs[10][1] - res_maxs[2][1]
        y = res_maxs[10][0] - res_maxs[2][0]
        degree = math.degrees(math.atan2(y, x))
        mat = cv2.getRotationMatrix2D(
            (res_maxs[1][1], res_maxs[1][0]), degree, 1.0)
        mks = cv2.warpAffine(mks, mat, (mks.shape[1], mks.shape[0]))
    return (mks)


# 表面の処理
def preprocessing(filepath):
    marker = cv2.imread('/content/drive/MyDrive/marker.png', 0)
    mks = cv2.imread(filepath, 0)
    h, w = mks.shape
    if h > w:
        mks = cv2.rotate(mks, cv2.ROTATE_90_CLOCKWISE)
    res_maxs = match(mks, marker)
    mks, res_maxs, back = rotate(mks, marker, res_maxs)
    mks = tilt(mks, res_maxs)
    return (mks)

# 両面の時の処理
def preprocessing_both(filepath1, filepath2):
    marker = cv2.imread('/content/drive/MyDrive/marker.png', 0)
    mks1 = cv2.imread(filepath1, 0)
    mks2 = cv2.imread(filepath2, 0)
    h, w = mks1.shape
    if h > w:
        mks1 = cv2.rotate(mks1, cv2.ROTATE_90_CLOCKWISE)
    h, w = mks2.shape
    if h > w:
        mks2 = cv2.rotate(mks2, cv2.ROTATE_90_CLOCKWISE)
    res_maxs1 = match(mks1, marker)
    res_maxs2 = match(mks2, marker)
    mks1, res_maxs1, back = rotate(mks1, marker, res_maxs1)
    if back:
        mks2 = cv2.rotate(mks2, cv2.ROTATE_180)
        res_maxs2 = match(mks2, marker)
    mks1 = tilt(mks1, res_maxs1)
    mks2 = tilt(mks2, res_maxs2, 1)
    return (mks1, mks2)

# マークエリア切り出し
def cut_markarea(mks):
    marker = cv2.imread('/content/drive/MyDrive/marker.png', 0)
    res_maxs = match(mks, marker)
    mark_areas = []
    for i in range(0, len(res_maxs), 2):
        y1, x1 = res_maxs[i]
        y2, x2 = res_maxs[i+1]
        if i == 0:
            mark_area = mks[y1: y2+marker.shape[0], x1: x2+marker.shape[1]]
        else:
            mark_area = mks[y1: y2, x1: x2]
        mark_areas.append(mark_area)
    return (mark_areas)


def exnum(mark_area):
    # 受験番号のマーク処理
    upper_margin = 2
    bottom_margin = 2
    row = 10
    total_col = 8
    total_row = row + upper_margin + bottom_margin

    # markareaの処理
    area = cv2.resize(mark_area, (total_col*100, total_row*100))
    area = area[upper_margin*100:(total_row-bottom_margin)*100]
    area = cv2.GaussianBlur(area, (5, 5), 0)
    res, area = cv2.threshold(area, 50, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    area = 255 - area

    # markの認識
    result = []
    for i in range(total_col):
        col_area = area[:, i*100: (i+1)*100]
        area_sum = []
        for s in range(row):
            onemark_area = col_area[s*100:(s+1)*100, :]
            onemark_area_sum = np.sum(onemark_area)
            area_sum.append(onemark_area_sum)
        result.append(area_sum)
    result = np.array(result)
    result_num = ''
    for i in range(len(result)):
        threshold = 450000
        while (len(np.where((result[i] > threshold) == True)[0]) < 1) and (threshold > 200000):
            threshold -= 20000
        else:
            if (threshold <= 200000) or (len(np.where((result[i] > threshold) == True)[0]) == 2):
                num = "*"
                result_num += num
            else:
                num = np.where((result[i] > threshold) == True)
                num = num[0]
                result_num += str(num[0])
    return (result_num)


def ansnum(mark_area):
    # 解答部分の処理
    row = 20
    upper_margin = 3
    bottom_margin = 1
    total_col = 9
    total_row = row + upper_margin + bottom_margin

    # markareaの処理
    area = cv2.resize(mark_area, (total_col*100, total_row*100))
    area = area[upper_margin*100:(total_row-bottom_margin)*100]
    area = cv2.GaussianBlur(area, (5, 5), 0)
    res, area = cv2.threshold(area, 50, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    area = 255 - area

    # markの認識
    result = []
    for i in range(row):
        row_area = area[i*100: (i+1)*100, :]
        area_sum = []
        for s in range(total_col):
            onemark_area = row_area[:, s*100:(s+1)*100]
            onemark_area_sum = np.sum(onemark_area)
            area_sum.append(onemark_area_sum)
        result.append(area_sum)
    result = np.array(result)
    result_num = []

    for i in range(len(result)):
        threshold = 450000
        while (len(np.where((result[i] > threshold) == True)[0]) < 1) and (threshold > 300000):
            threshold -= 20000
        else:
            if (threshold <= 300000) or (len(np.where((result[i] > threshold) == True)[0]) == 2):
                num = []
            else:
                num = np.where((result[i] > threshold) == True)
                num = num[0] + 1
                num = num.tolist()
        if num == []:
            result_num.append(0)
        else:
            result_num.append(num[0])
    return (result_num)


def getnumbers(mark_areas):
    answers = []
    for i in range(len(mark_areas)):
        if i == 0:
            mark_area = mark_areas[i]
            numbers = exnum(mark_area)
        else:
            answers += (ansnum(mark_areas[i]))
    return (numbers, answers)


def getnumbers_both(mark_areas1, mark_areas2):
    answers = []
    for i in range(len(mark_areas1)):
        if i == 0:
            numbers = exnum(mark_areas1[i])
        else:
            answers += (ansnum(mark_areas1[i]))

    for i in range(len(mark_areas2)):
        answers += (ansnum(mark_areas2[i]))
    return (numbers, answers)


# excelにデータを書き込む
def excel_process(numbers_list, answers_list, wb, row_fin, col):
    nonumbers = []
    noanswers = []
    nonums = []
    for i in range(len(numbers_list)):
        for sheetname in wb.sheetnames:
            if numbers_list[i] in sheetname:
                ws = wb[sheetname]
                for s in range(3, row_fin):
                    ws.cell(row=s, column=col, value=answers_list[i][s-3])
                break
        else:
            nonumbers.append(numbers_list[i])
            noanswers.append(answers_list[i])
            nonums.append(i)
    return (nonumbers, noanswers, nonums)


# 生徒番号のない生徒の処理
def nonumber(nonumbers, nosheets_num, filepaths):
    newnums = []
    print("以下の受験番号の生徒は存在しません")
    for i in range(len(nonumbers)):
        print("---------------------------------------------------------------------------")
        print(nonumbers[i]+"は存在しません。次の解答用紙を確認してください。")
        print(filepaths[nosheets_num[i]].split("/")[-1])
        newnums.append(input("正しい生徒番号を入力してください"))
    return (newnums)


def single_scan(filepaths, col, row_fin):
    numbers_list = []
    answers_list = []
    mks_list = []
    for filepath in filepaths:
        print(filepath)
        mks = preprocessing(filepath)
        mark_areas = cut_markarea(mks)
        numbers, answers = getnumbers(mark_areas)
        numbers_list.append(numbers)
        answers_list.append(answers)
        mks_list.append(mks)

    path = '/content/drive/MyDrive/sendaitest/answer_data.xlsx'
    wb = openpyxl.load_workbook(path)
    nonumbers, noanswers, nonums = excel_process(
        numbers_list, answers_list, wb, row_fin, col)
    while len(nonumbers) > 0:
        newnums = nonumber(nonumbers, nonums, filepaths)
        nonumbers, noanswers, nonums = excel_process(
            newnums, noanswers, wb, row_fin, col)
    else:
        wb.save(path)
        print("----------------------------------------------------------------------------------------")
        print("処理は完了しました")


def double_scan(filepaths, col, row_fin):
    numbers_list = []
    answers_list = []

    for i in range(0, len(filepaths), 2):
        print(filepaths[i], filepaths[i+1])
        mks1, mks2 = preprocessing_both(filepaths[i], filepaths[i+1])
        mark_areas1 = cut_markarea(mks1)
        mark_areas2 = cut_markarea(mks2)
        numbers, answers = getnumbers_both(mark_areas1, mark_areas2)
        numbers_list.append(numbers)
        answers_list.append(answers)

    path = '/content/drive/MyDrive/sendaitest/answer_data.xlsx'
    wb = openpyxl.load_workbook(path)
    nonumbers, noanswers, nonums = excel_process(
        numbers_list, answers_list, wb, row_fin, col)
    while len(nonumbers) > 0:
        newnums = nonumber(nonumbers, nonums, filepaths)
        nonumbers, noanswers, nonums = excel_process(
            newnums, noanswers, wb, row_fin, col)
    else:
        wb.save(path)
        print("----------------------------------------------------------------------------------------")
        print("処理は完了しました")

